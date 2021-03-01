#functions to run in bought tab
import tkinter as tk
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import numbers
from os import path
    
def make_output_box(root):
    output=tk.Text(root,width=75,height=4,wrap=tk.WORD,bg='white')
    output.pack(side=tk.BOTTOM)

    return output

def load_workbook(wb_path):
    #function loads a workbook, if the wb doesnt exist it crteates one
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    #Create and initialize workbook with headers
    create_workbook(wb_path)
    return openpyxl.load_workbook(wb_path)
    
def create_workbook(wb_path):
    #create inventory workbook
    wb=openpyxl.Workbook()
    ws1=wb.active
    ws1.title="Stock"

    #initialize advanced stock headers
    stock_headers=["Name","Price","Size","Brand","Stock","ID"]
    for i,cell in enumerate(ws1["A1":"F1"][0]): #indexed at 0 because returns cells in a tuple of a size 1 tuples
        cell.value=stock_headers[i]
    
    
    ws2=wb.create_sheet("History")
    #initialize history headers
    history_headers=["Name","Price(Bought)","Size","Brand","Price(Sold)","Profit","Total Profit"]
    for i,cell in enumerate(ws2["A1":"G1"][0]):
        cell.value=history_headers[i]


    wb.save(wb_path)

def stock_to_update(name_adding,ws):
    names_in_stock=list(map(lambda x: x.value,ws['A']))[1:] #gets names of all sneakers in stock
    #index in name_in_stock corresponds to row of index+2, i.e if index is 3 then in spreadsheet that sneaker will be in row 
    return [i+2 for i in range(len(names_in_stock)) if names_in_stock[i]==name_adding]

def add_to_stock(info_list,wb_path): #can probs just add straight to a_stock
   
    wb=load_workbook(wb_path)
    stock_sheet=wb['Stock']
    last_row=stock_sheet.max_row
    if "$" in info_list[1]:
        info_list[1]=float(info_list[1][1:])#remove $ sign from price and make number float
    else:
        info_list[1]=float(info_list[1])
    
    #if stock empty add sneaker with id of 1
    if last_row==1:
        #add info to cells
        for i,cell in enumerate(stock_sheet["A2":"D2"][0]):
            if i==1: #price column
                cell.value=info_list[1]
                cell.number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE #format cell as currency
                continue
            cell.value=info_list[i]

        #add stock of 1 and ID as 1
        stock_sheet["E2"]=1
        stock_sheet["F2"]=1
    else:
        #if sheet already has entries
        for i,cell in enumerate(stock_sheet["A"+str(last_row+1):"D"+str(last_row+1)][0]):

            if i==1: #price column
                cell.value=info_list[1]
                cell.number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE #format cell as currency
            cell.value=info_list[i]

        rows_to_update=stock_to_update(info_list[0],stock_sheet)
        for row in rows_to_update:
            stock_sheet["E"+str(row)]=len(rows_to_update)

        stock_sheet["E"+str(last_row+1)]=len(rows_to_update) #stock num
        stock_sheet["F"+str(last_row+1)]=stock_sheet["F"+str(last_row)].value+1 #ID_num


    wb.save(wb_path)