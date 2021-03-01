import tkinter as tk
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import numbers
from os import path


#gets row numbers of sneaker if given name is subset of sneaker name
def get_potetntial_rows(name_selling,ws):
    names_in_stock=list(map(lambda x: x.value,ws['A']))[1:] #gets names of all sneakers in stock

    return [i+2 for i in range(len(names_in_stock)) if name_selling in names_in_stock[i]] #index in name_in_stock corresponds to row of index+2

def get_options(name_selling,stock_sheet):
    #make list to hold potential sneakers to sell based on name search (name,price,size,id)
    possible_s=["Which one: \n"]

    rows_to_stringify=get_potetntial_rows(name_selling,stock_sheet)
    for row in rows_to_stringify:
        possible_s.append([f'Name: {stock_sheet["A"+str(row)].value} Price: {stock_sheet["B"+str(row)].value} Size: {stock_sheet["C"+str(row)].value} ID: {stock_sheet["F"+str(row)].value} \n'])
    
    return possible_s

#function to make box with list of options of shoes to sell based on name entered
def options_box(tab):
    box=tk.Text(tab,width=75,height=4,wrap=tk.WORD,bg='white')
    box.pack(side=tk.BOTTOM)  


def remove_dollar(price):
    #price is a string
    new_price=""
    for char in price:
        if char!='$':
            new_price+=char
    return new_price

def get_stock_sheet():
    if not path.exists("./docs/inventory.xlsx"):
        raise FileNotFoundError("Trying to access spreadsheet that doesn't exist")

    wb=openpyxl.load_workbook("./docs/inventory.xlsx")

    stock_sheet=wb['Stock']
    return stock_sheet

def sell_sneaker(id_to_sell,price_sold,wb_path):
    if not path.exists(wb_path):
        raise FileNotFoundError
    
    
    wb=openpyxl.load_workbook(wb_path)

    stock_sheet=wb['Stock']
    history_sheet=wb['History']
    last_srow=stock_sheet.max_row
    last_hrow=history_sheet.max_row #last row in history

    if id_to_sell<1:
        raise ValueError("Invalid ID")
    
    #find row to delete from stock
    row_to_delete=-1
    for c in stock_sheet["F"]:
        if c.value==id_to_sell:
            row_to_delete=c.row
            break

    if row_to_delete==-1:
        raise ValueError("ID not found in spreadsheet")

    letters=["A","B","C","D"]
    info_history=[ stock_sheet[ letters[i]+str(row_to_delete) ].value  for i in range(len(letters))]

    #delete sneaker from stock
    stock_sheet.delete_rows(row_to_delete)


    #add sneaker to history
    for i,cell in enumerate(history_sheet["A"+str(last_hrow+1):"D"+str(last_hrow+1)][0]): #gets cells in row last_hrow +1 from A to D column
        cell.value=info_history[i]
        if i==1: #price bought column
            cell.number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE

    if "$" in price_sold:
        price_sold=price_sold[1:] #remove $ sign from  price 
    price_sold=float(price_sold) #make it a float
    

    history_sheet["E"+str(last_hrow+1)]=price_sold#price sold shoe for
    history_sheet["E"+str(last_hrow+1)].number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE #format cell as currency
    
    history_sheet["F"+str(last_hrow+1)]="=SUM(E"+str(last_hrow+1)+"-B"+str(last_hrow+1)+")" #profit = price sold - price bought
    history_sheet["F"+str(last_hrow+1)].number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE

    history_sheet["G2"]="=SUM(F2:F"+str(last_hrow+1)+")" #cell holdiing totalProfit
    history_sheet["G2"].number_format =numbers.FORMAT_CURRENCY_USD_SIMPLE

    wb.save(wb_path)